#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''
Created on 2017年7月3日
@author: Irony."[讽刺]
@site: alyl.vip, orzorz.vip, irony.coding.me , irony.iask.in , mzone.iask.in
@email: 892768447@qq.com
@file: AutoReportUtil
@description: 
'''
import binascii
from datetime import datetime, timedelta
from email.header import Header
from imp import reload
import json
import logging
import os
import re
from smtplib import SMTP
import traceback
from urllib.parse import quote

from jinja2 import Template
import jpype
from openpyxl.reader.excel import load_workbook
import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder, \
    MultipartEncoderMonitor
import yaml

import AutoReportGlobals  # @UnresolvedImport


# from datetime import datetime
__version__ = "0.0.1"

yaml.reader.Reader.NON_PRINTABLE = re.compile(
    u'[^\x09\x0A\x0D\x20-\x7E\x85\xA0-\uD7FF\uE000-\uFFFD\U00010000-\U0010FFFF]')


class WorksheetToImage:

    def __init__(self, dataDir, fileout, areas):
        self.dataDir = dataDir
        self.fileout = fileout
        self.areas = areas

        License = jpype.JClass("com.aspose.cells.License")
        l = License()
        l.setLicense("datas/lib/license.xml")

        self.Workbook = jpype.JClass("com.aspose.cells.Workbook")
        self.ImageFormat = jpype.JClass("com.aspose.cells.ImageFormat")
        self.ImageOrPrintOptions = jpype.JClass(
            "com.aspose.cells.ImageOrPrintOptions")
        self.SheetRender = jpype.JClass("com.aspose.cells.SheetRender")

    @classmethod
    def start(self, logger, dataDir, fileout, areas):
        '''
        :param dataDir: 输出图片目录
        :param filein: 临时渲染模版文件
        :param areas: 截图区域
        '''
        # 库
#         jars = "datas/lib/bcprov-jdk16-146.jar;datas/lib/aspose-cells-17.6.jar"
        jars = "datas/lib/bcprov-jdk16-146.jar;datas/lib/excel.jar"
        logger.debug("start convert excel to image")
        logger.debug("image save dir: {0}".format(dataDir))
        logger.debug("template file is: {0}".format(fileout))
        logger.debug("capture image areas: {0}".format(areas))
        # 先判断jvm是否启动(这是官网的一个BUG)
        AutoReportGlobals.JVMStarted = jpype.isJVMStarted()
#         if not jpype.isJVMStarted():
        jpype.startJVM(jpype.getDefaultJVMPath(),
                       "-ea", "-Xmn128m", "-Xms512M", "-Xmx512M",
                       "-Djava.class.path={0}".format(jars)
                       )
        jpype.attachThreadToJVM()  # 解决多线程问题
        wst = WorksheetToImage(dataDir, fileout, areas)
        wst.main()
        jpype.shutdownJVM()  # 不要关闭否则第二次无法启动
        logger.debug("end convert excel to image")

    def main(self):
        imageFormat = self.ImageFormat

        # Instantiate a workbook with path to an Excel file
        book = self.Workbook(self.fileout)

        # 保存为html文件
        # book.save("data/out.html",12)

        # Create an object for ImageOptions
        imgOptions = self.ImageOrPrintOptions()
#         print(imgOptions.getQuality())
#         imgOptions.setQuality(100)
        imgOptions.setOnePagePerSheet(True)

        # Set the image type
        # 输出图片格式
#         imgOptions.setImageFormat(imageFormat.getJpeg())
        imgOptions.setImageFormat(imageFormat.getPng())

        CalculationOptions = jpype.JClass(
            "com.aspose.cells.CalculationOptions")
        opt = CalculationOptions()

        for sheetname, areas in self.areas.items():
            # Get the first worksheet.
            # 获取sheet
            sheet = book.getWorksheets().get(sheetname)
#             print("sheet ", sheetname)
            # 对sheet中的公式进行计算
            sheet.calculateFormula(opt, True)

            # 设置区域
            pageSetup = sheet.getPageSetup()
            # 去掉边距
            pageSetup.setBottomMargin(0.)
            pageSetup.setLeftMargin(0.)
            pageSetup.setRightMargin(0.)
            pageSetup.setTopMargin(0.)

            # 区域
            for index, area in enumerate(areas):
                #                 print("area: ", area)
                pageSetup.setPrintArea(area)
        #         pageSetup.setPrintTitleRows("$2:$2")

                # Create a SheetRender object for the target sheet
                sr = self.SheetRender(sheet, imgOptions)
                for page in range(sr.getPageCount()):
                    # Generate an image for the worksheet
                    if index > 8:  # 比如index = 9
                        index = index * 10  # 9*10 = 90  +1 ==91,92,93
                    sr.toImage(
                        page, os.path.join(self.dataDir, "%d.png" % (index + 1)))
        # Print message
#         print("Images generated successfully.")


MailTemplate = Template("""<html>
    <body>
        <center><h1>{{ subject }}</h1></center><br />
        {% for ipath in imglist %}
        <h2>图片：{{ ipath }}</h2><br />
        <img src="cid:{{ ipath }}" alt="{{ ipath }}"><br />
        {% endfor %}
        <br />
        #任务ID：{{ eid }}<br />
        #说明：回复命令格式：<br />
        ## 发送+任务ID,1,3,5 表示只发送135序号的图片(不含+号)<br />
        ## 发送+任务ID       表示全部发送(不含+号)<br />
    <body>
</html>""")


def makeId():
    while 1:
        eid = binascii.hexlify(os.urandom(3)).decode()
        if not AutoReportGlobals.EmailIds or \
                eid not in AutoReportGlobals.EmailIds:
            return eid


def sendEmailTextByConfig(file, logger, subject, text):
    wb = load_workbook(file)
    # 预发送人
    preSheet = wb["预发送"]
    if not preSheet:
        return logger.error("无法找到预发送人表")
    preSend = [str(cell.value) for cell in preSheet["A"] if str(
        cell.value) != "None"]  # list(preSheet.columns)[0]]
    logger.debug("预发送人: {0}".format(preSend))
    sendEmailText(preSend, subject, text)


def sendEmailText(to, subject, text):
    if not to:
        return logging.getLogger("AutoReportMms").debug("邮件接收人为空")
    config = json.load(open("datas/config.json", "r"))
    From = config.get("USERNAME", "")
    Pwd = config.get("PASSWORD", "")
    FROM = From
    BODY = """From: {0}
To: {1}
Subject: {2}

{3}
""".format(From, ",".join(to), subject, text).encode("gb2312")
    try:
        smtp = SMTP("smtp.139.com", 25)
        smtp.login(From, Pwd)
        smtp.sendmail(
            FROM, to, BODY
        )
        smtp.quit()
        return True
    except Exception as e:
        logging.getLogger("AutoReportMms").error(str(e))
    return False


def sendEmail(subject, message, preSend):
    '''发送邮件
    :param subject: 邮件主题
    :param message: 邮件
    :param preSend: 发送人
    '''
    if not preSend:
        return False
#     print(type(preSend),preSend)
    config = json.load(open("datas/config.json", "r"))
    From = config.get("USERNAME", "")
    Pwd = config.get("PASSWORD", "")
    message["Subject"] = Header(subject, "gb2312")  # 邮件主题
    message["From"] = From  # 发件人邮箱
    TO = [people + "@139.com" for people in preSend]
    message["To"] = ",".join(TO)  # 接收人列表
    try:
        smtp = SMTP("smtp.139.com", 25)
        smtp.login(From, Pwd)  # 登录
        smtp.sendmail(  # 发送邮件
            message["From"], TO, message.as_string()
        )
        smtp.quit()
        return True
    except Exception as e:
        logging.getLogger("AutoReportMms").error(str(e))
    return False


def sendTextMms(subject, content, peoples):  # 发送文本彩信
    '''
    #发送文本彩信
    :param subject: 主题
    :param content: 内容
    :param peoples: 接收人
    '''
    fields = [
        ("PhoneNumber", ",".join(peoples)),  # 发送对象
        ("MMSFrom", ""),  # 发送人
        ("MMSSubject", subject),  # 发送主题
        ("MMSText", content),  # 文字内容
        ("MMSBCC", "No"),
        ("MMSDeliveryReport", "No"),
        ("MMSReadReport", "No"),
        ("MMSPriority", "Normal"),
        ("MMSMessageClass", "Personal"),
        ("MMSForwardLock", "No"),
        ("DRMRestrict", "No"),
        ("DRMRestrictTextXML", ""),
        #       (  "DRMPermissionPlay",""),
        ("DRMConstraintCount", ""),
        ("DRMConstraintStart", ""),
        ("DRMConstraintEnd", ""),
        ("DRMConstraintInterval", ""),
        ("MMSHEADERS", ""),
        ("Submit", "Submit")
    ]
    encoder = MultipartEncoder(
        fields=fields
    )

    # print(fields + _fields)

    monitor = MultipartEncoderMonitor(encoder)

    req = requests.post(
        "http://10.105.160.68:8800/Send%20MMS%20Message.htm",
        data=monitor,
        timeout=120,
        headers={
            "Accept": "image/gif, image/jpeg, image/pjpeg, application/x-ms-application, application/xaml+xml, application/x-ms-xbap, application/vnd.ms-excel, application/vnd.ms-powerpoint, application/msword, */*",
            "Cache-Control": "no-cache",
            "Connection": "Keep-Alive",
            "Referer": "http://10.105.160.68:8800/Send%20MMS%20Message.htm",
            "Cookie": "ASP.NET_SessionId=igwzr5gsgeyssz11slb0pyuq",
            "Content-Type": monitor.content_type,
            "User-Agent": "Auto SMS BY: 892768447 Version: 1.0"
        }
    )
    return req.text.find("Message Submitted") > -1


def sendMms(subject, dataDir, peoples, whichs, text=""):  # 发送彩信
    '''
    #发送彩信
    :param subject: 主题
    :param dataDir: 图片目录
    :param peoples: 接收对象
    :param whichs: 哪些需要发送
    :param text: 附加内容
    '''
    fields = [
        ("PhoneNumber", ",".join(peoples)),  # 发送对象
        ("MMSFrom", ""),  # 发送人
        ("MMSSubject", subject),  # 发送主题
        ("MMSText", text),  # 文字内容
        ("MMSBCC", "No"),
        ("MMSDeliveryReport", "No"),
        ("MMSReadReport", "No"),
        ("MMSPriority", "Normal"),
        ("MMSMessageClass", "Personal"),
        ("MMSForwardLock", "No"),
        ("DRMRestrict", "No"),
        ("DRMRestrictTextXML", ""),
        #       (  "DRMPermissionPlay",""),
        ("DRMConstraintCount", ""),
        ("DRMConstraintStart", ""),
        ("DRMConstraintEnd", ""),
        ("DRMConstraintInterval", ""),
        ("MMSHEADERS", ""),
        ("Submit", "Submit")
    ]
    _fields = []
    if len(whichs) == 0:  # 全部
        whichs = [p[:-4] for p in os.listdir(dataDir)]
    for which in whichs:
        try:
            int(which)
        except:
            continue
        path = os.path.join(dataDir, str(which) + ".png")
        _fields.append(("MMSFile", (
            quote(os.path.basename(path)),  # 获取文件名并对文件名编码
            open(path, "rb"),  # 二进制流
            "image/x-png")
        ))
    if not _fields:
        return -1
    encoder = MultipartEncoder(
        fields=fields + _fields
    )

    # print(fields + _fields)

    monitor = MultipartEncoderMonitor(encoder)

    req = requests.post(
        "http://10.105.160.68:8800/Send%20MMS%20Message.htm",
        data=monitor,
        timeout=120,
        headers={
            "Accept": "image/gif, image/jpeg, image/pjpeg, application/x-ms-application, application/xaml+xml, application/x-ms-xbap, application/vnd.ms-excel, application/vnd.ms-powerpoint, application/msword, */*",
            "Cache-Control": "no-cache",
            "Connection": "Keep-Alive",
            "Referer": "http://10.105.160.68:8800/Send%20MMS%20Message.htm",
            "Cookie": "ASP.NET_SessionId=igwzr5gsgeyssz11slb0pyuq",
            "Content-Type": monitor.content_type,
            "User-Agent": "Auto SMS BY: 892768447 Version: 1.0"
        }
    )
    return req.text.find("Message Submitted") > -1


def initLogger(path, LogName):
    # 日志文件
    os.makedirs(os.path.dirname(path), exist_ok=True)
#     handler = logging.FileHandler(path + "/" + LogName + "-" +
# datetime.now().strftime("%Y-%m-%d") + ".log", mode="w", encoding="gbk")
    handler = logging.FileHandler(path + ".log", mode="w", encoding="gbk")
    formatter = logging.Formatter(
        "[%(asctime)s %(name)s:%(lineno)s] P:%(process)d T:%(thread)d %(levelname)-8s %(message)s<br />")
    handler.setFormatter(formatter)

    logger = logging.getLogger(LogName)
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG)
    return logger


# def initJobs(logger):
#     for name in os.listdir("jobs"):
# #         if not name.endswith(".py") or not name.endswith("Job.py"):
# #             continue
#         path = "jobs/" + name
#         if name.endswith("Job") and os.path.isdir(path):
#             jname = "job"  # name[:-3]
#             module = __import__("jobs.{0}.{1}".format(name, jname), fromlist=(jname,))
#             reload(module)
# #             print(module, dir(module))
#             logger.debug("find module jobs.{0}.{1}".format(
#                 name, jname, getattr(module, "Enable", False)))
#             if getattr(module, "Enable", False):
#                 logger.debug("load module jobs.{0}.{1}".format(name, jname))
#                 try:
#                     module.init()
#                 except Exception as e:
#                     traceback.print_exc()
#                     logger.warn(
#                         "load module {0}{1} error: {2}".format("jobs.", jname, e))

def initJobs(logger):
    for path, _, files in os.walk("jobs"):
        for name in files:
            _path = os.path.join(path, name)
            if name == "job.py":
                _module = _path[:-3].replace("../", "").replace("\\", ".")
                module = __import__(_module, fromlist=("job",))
                reload(module)
#                 logger.debug("find module {0},enabled: {1}".format(
#                     module, getattr(module, "Enable", False)))
                if getattr(module, "Enable", False):
                    logger.debug("load module {0}".format(_module))
                    try:
                        module.init()
                    except Exception as e:
                        traceback.print_exc()
                        logger.warn(
                            "load module {0} error: {2}".format(str(_module), str(e)))


def getConfig(tname):
    logger = logging.getLogger("AutoReportMms")
    logger.debug("tname: " + tname)
    wb = load_workbook(tname)
    sheet = wb["配置"]
    config = yaml.load(sheet.cell(row=1, column=1).value)
    kwargs = {}
    # 解析时间配置
    cron = config.get("cron", {})  # cron风格
    logger.debug("cron:" + str(cron))
    # cron
    if cron and cron.get("启用", False):
        trigger = "cron"
        year = cron.get("年", None)
        if year:
            kwargs["year"] = year
        month = cron.get("月", None)
        if month:
            kwargs["month"] = month
        day = cron.get("日", None)
        if day:
            kwargs["day"] = day
        week = cron.get("周", None)
        if week:
            kwargs["week"] = week
        day_of_week = cron.get("日周", None)
        if day_of_week:
            kwargs["day_of_week"] = day_of_week
        hour = cron.get("时", None)
        if hour:
            kwargs["hour"] = hour
        minute = cron.get("分", None)
        if minute:
            kwargs["minute"] = minute
        second = cron.get("秒", None)
        if second:
            kwargs["second"] = second
        start_date = cron.get("开始日期", None)
        if start_date:
            kwargs["start_date"] = start_date
        end_date = cron.get("结束日期", None)
        if end_date:
            kwargs["end_date"] = end_date
        return trigger, kwargs, config

    time = config.get("时间")
    logger.debug("时间:" + str(time))
    start_date = time.get("开始日期", 0)
    end_date = time.get("结束日期", 0)
    run_date = time.get("执行一次", 0)

    # 周期执行
    if start_date:
        trigger = "interval"
        weeks = time.get("周", 0)
        days = time.get("天", 0)
        hours = time.get("时", 0)
        minutes = time.get("分", 0)
        seconds = time.get("秒", 0)
        if weeks == days == hours == minutes == seconds == 0:
            raise Exception("周天时分秒不能全部为0")
        if weeks:
            kwargs["weeks"] = weeks
        if days:
            kwargs["days"] = days
        if hours:
            kwargs["hours"] = hours
        if minutes:
            kwargs["minutes"] = minutes
        if seconds:
            kwargs["seconds"] = seconds
        if start_date:
            kwargs["start_date"] = start_date
        if end_date:
            kwargs["end_date"] = end_date
    elif run_date:
        trigger = "date"
        kwargs["run_date"] = run_date
    else:
        raise Exception("执行时间配置错误")
    logger.debug(
        "trigger:{0}, kwargs:{1}, config:{2}".format(trigger, kwargs, config))
    return trigger, kwargs, config


def CURRENT_TIME(days=0, seconds=0, microseconds=0):
    '''获取当前日期时间'''
    try:
        return datetime.now() + timedelta(days=days, seconds=seconds, microseconds=microseconds)
    except:
        return datetime.now()


def CURRENT_MONTH(month=0):
    '''获取当前月'''
    now = datetime.now()
    try:
        return datetime(now.year, now.month + month, 1)
    except:
        return now


def DATE_FORMAT(date, formats="%Y%m%d"):
    '''格式化当前时间'''
    try:
        return date.strftime(formats)
    except:
        return date.strftime("%Y%m%d")

if __name__ == "__main__":
    print(sendEmailText(['18228111461@qq.com'],'test','test'))