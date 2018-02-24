#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''
Created on 2017年7月3日
@author: Irony."[讽刺]
@site: alyl.vip, orzorz.vip, irony.coding.me , irony.iask.in , mzone.iask.in
@email: 892768447@qq.com
@file: jobs.CommonLibs.MmsDaily
@description: 彩信日报发送公共库
'''
from datetime import datetime
from email.header import Header
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
from time import localtime, strftime

from jinja2 import Template
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border, Side
import requests
from sqlalchemy.engine import create_engine

from AutoReportUtil import WorksheetToImage, sendEmail, makeId  # @UnresolvedImport
# @UnresolvedImport
from AutoReportUtil import initLogger, DATE_FORMAT, CURRENT_TIME, CURRENT_MONTH


# import AutoReportGlobals  # @UnresolvedImport
__version__ = "0.0.1"

URL = "http://10.105.4.50:58080/BingoInsight/dataservice.dsr"

HEADERS = {
    "Accept": "*/*",
    "Referer": "http://10.105.4.50:58080/BingoInsight/portal/ClientBin/Bingosoft.DataOne.Portals.Boot.xap",
    "Content-Type": "text/json;chartset=utf-8",
    "User-Agent": "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729; InfoPath.2)",
    "Connection": "Keep-Alive",
    "Cache-Control": "no-cache",
    "Cookie": "JSESSIONID=D639101A895F1F7DC61011877347D8AC",
}

BORDER = Border(
    left=Side("thin", "FF000000"),
    right=Side("thin", "FF000000"),
    top=Side("thin", "FF000000"),
    bottom=Side("thin", "FF000000"),
    vertical=Side("thin", "FF000000"),
    horizontal=Side("thin", "FF000000")
)

MailTemplate = Template("""<html>
    <body>
        <center><h1>{{ subject }}</h1></center><br />
        {% for ipath in imglist %}
        <h2>图片：{{ ipath }}</h2><br />
        <img src="cid:{{ ipath }}" alt="{{ ipath }}"><br />
        {% endfor %}
        <br />
        #说明 任务ID：{{ eid }}<br />
        #回复命令格式：(注意回复不要附带原始邮件内容)<br /><br />
        ##序号发送：<br />
        M{{ eid }},1,3,5<br /><br />
        ##全部发送：<br />
        M{{ eid }}<br />
    <body>
</html>""")


class MmsDaily:

    def __init__(self, Id, config, path, subject, port=8989):
        os.makedirs("datas/logs/", exist_ok=True)
        os.makedirs("datas/tmps/", exist_ok=True)
        self.logger = initLogger("datas/logs/" + Id, Id)
        self.config = config
        self.subject = subject
        self.port = port
        filename = os.path.basename(path)  # 文件名
        self.filein = path
        self.fileout = os.path.join("datas/tmps/" + filename)
        self.dataDir = os.path.join(
            "datas/tmps/", os.path.basename(os.path.splitext(self.fileout)[0]))
        os.makedirs(self.dataDir, exist_ok=True)

    def start(self):
        self.logger.debug("pid: %s" % os.getpid())
        # 加载模版文件开始任务
        wb = load_workbook(self.filein)
        # 解析配置
        self.parseConfig(wb)
        # 预发送邮件
        if not self.preSend:
            return self.logger.warn("预发送人员列表为空")
        # 填充数据
        self.fillData(wb)
        # 修改截图区域的边框
        self.modifyBorder(wb, self.captureArea)
        # 删除多余sheet
        try:
            wb.remove(wb["配置"])
            wb.remove(wb["预发送"])
            wb.remove(wb["联系人"])
        except:
            pass
        # 保存填充后的excel
        wb.save(self.fileout)
        # 指定区域截图
        WorksheetToImage.start(
            self.logger, self.dataDir, self.fileout, self.captureArea)
        self.logger.info("production task success")

        eid = makeId()  # 获取唯一编码
        einfo = {
            "typ": 0,  # 储存方式,0=预发送,1=正式发送(邮件方式)
            "subject": self.subject,  # 储存主题信息
            "dataDir": self.dataDir,  # 储存数据目录
            "preSend": self.preSend,  # 储存预发送人
            "formalSend": self.formalSend  # 储存正式发送人
        }
        # 此处改为post本地提交
#         AutoReportGlobals.EmailIds[eid] = einfo
        # 此处改为post本地提交
        _json = {
            "type": 0,  # 储存id
            "eid": eid,
            "einfo": einfo
        }
        req = requests.post("http://127.0.0.1:%s/api/email" %
                            self.port, json=_json)
        if req.json().get("code") != 1:
            return self.logger.error("储存生成的email id 错误")
        self.logger.debug("储存生成的email id 成功")
#         print("sqdjob: ", AutoReportGlobals.EmailIds)
        self._sendEmail(eid, self.subject, self.dataDir,
                        self.preSend, self.formalSend, 0)
        self.logger.info("send pre email success")

    def _sendEmail(self, eid, subject, dataDir, preSend, formalSend, typ=0):
        peoples = formalSend if typ else preSend  # 如果采用邮箱正式发送
        subject = subject + datetime.now().strftime(" - %Y%m%d %H:%M:%S")
        subject = subject if typ else subject + " - 预览"
        message = MIMEMultipart()
        message["Subject"] = Header(subject, "gb2312")  # 邮件主题
        # 邮件正文
        imglist = os.listdir(dataDir)  # 遍历图片
        imglist.sort(key=lambda x: int(x[:-4]))  # 排序
        content = MIMEText(MailTemplate.render(
            subject=subject, imglist=imglist, eid=eid
        ), _subtype="html", _charset="gb2312")
        message.attach(content)
        # 添加图片附件
        for ipath in imglist:
            with open(os.path.join(dataDir, ipath), "rb") as fp:
                img = MIMEImage(fp.read())
                img.add_header("Content-ID", ipath)
                message.attach(img)
        sendEmail(subject, message, peoples)
        # 发送微信图片
#         if not AutoReportGlobals.WeixinBot: return
#         for ipath in imglist:
#             AutoReportGlobals.WeixinBot.sendImage(os.path.join(dataDir, ipath))
        # 调整为多进程网页请求
        _json = {
            "type": "image",
            "data": [os.path.abspath(os.path.join(dataDir, ipath)) for ipath in imglist]
        }
#         print(_json.get("data"))
        try:
            req = requests.post(
                "http://127.0.0.1:%s/api/wechat/send" % self.port, json=_json)
#             print(req.text)
            if req.json().get("code") != 1:
                return self.logger.error("微信群消息发送失败")
            self.logger.debug("微信群图片发送完毕")
            _json = {
                "type": "text",
                "data": """任务ID：%s\r\n
说明：回复命令格式：\r\n
日报发送,%s,1,2,3,4,5,6,7,8,9 表示只发送指定序号的图片\r\n
日报发送,%s 表示全部发送""" % (eid, eid, eid)
            }
            req = requests.post(
                "http://127.0.0.1:%s/api/wechat/send" % self.port, json=_json)
#             print(req.text)
            if req.json().get("code") != 1:
                return self.logger.error("微信群消息发送失败")
            self.logger.debug("微信群文字消息发送完毕")
        except Exception as e:
            self.logger.error(str(e))

    def parseConfig(self, wb):
        # 截图区域配置
        self.logger.debug("start parse config")
        self.captureArea = self.config.get("截图", {})
        # 发送人配置
        # 预发送人
        preSheet = wb["预发送"]
        if not preSheet:
            raise Exception("无法找到预发送人表")
        self.preSend = [str(cell.value) for cell in preSheet["A"] if str(
            cell.value) != "None"]  # list(preSheet.columns)[0]]
        self.logger.debug("预发送人: {0}".format(self.preSend))
        # 获取正式发送人
        formalSheet = wb["联系人"]
        if not formalSheet:
            raise Exception("无法找到联系人表")
        self.formalSend = [str(cell.value) for cell in formalSheet["A"] if str(
            cell.value) != "None"]  # list(formalSheet.columns)[0]]
        self.logger.debug("正式发送人: {0}".format(self.formalSend))
        '''
        self.preSend = [str(ph) for ph in sendPeople.get("预发送", []) if len(str(ph)) == 11]
        self.logger.debug("预发送人: {0}".format(self.preSend))
        # 获取正式发送人
        self.formalSend = sendPeople.get("正式发送", [])
        if self.formalSend:
            sheet, sarea, earea = self.formalSend.split(":")
            sheet = wb[sheet]
            self.formalSend = [str(cell.value) for row in sheet[sarea:earea]
                               for cell in row]
        self.preSend = [str(ph) for ph in self.formalSend if len(str(ph)) == 11]
        self.logger.debug("正式发送人: {0}".format(self.formalSend))
        '''
        # 填充区域配置
        self.fillArea = self.config.get("填充", [])
        self.logger.debug("end parse config")

    def fillData(self, wb):
        self.logger.debug("start fill data")
#         for sheetName, items in self.fillArea.items():
        for item in self.fillArea:
            self.fillItem(wb, item)
        self.logger.debug("end fill data")

    def fillItem(self, wb, item):
        sheetName = item.get("表", None)
        self.logger.warn("item sheetname is none: {0}".format(item))
        if not sheetName:
            return
        sheet = wb[sheetName]  # 获取sheet
        row = item.get("行", 1)
        col = item.get("列", 1)
        typ = item.get("类型", "")
        dbsrc = item.get("数据源", "")
        query = item.get("语句", "")
        self.logger.debug("sheetName:{0}\n"
                          "row:{1}\n"
                          "col:{2}\n"
                          "type:{3}\n"
                          "dbsrc:{4}\n".format(sheetName,
                                               row, col, typ, dbsrc
                                               )
                          )
        if not query:
            return self.logger.warn("查询语句不能为空")
        self.logger.debug("start get {0} data".format(sheetName))
        if typ == "selfsql":
            # 自助报表
            times, datas = self.queryWeb(dbsrc, query)
        else:
            # 连接数据查询
            times, datas = set(), self.queryDb(typ, query)
        self.logger.debug("end get {0} data".format(sheetName))

        self.logger.debug("start fill {0} data".format(sheetName))
        for r, item in enumerate(datas):
            for c, value in enumerate(item):
                if isinstance(value, str) and 1 <= len(value) <= 8:
                    try:
                        value = float(value)  # 尝试转成float类型
                    except:
                        pass
                try:  # 时间戳转换
                    if c in times:
                        value = localtime(int(str(value)[:10]))
                        value = strftime("%Y-%m-%d %H:%M:%S", value)
                except:
                    pass
                sheet.cell(row=row + r, column=col + c).value = value
#                     cell = sheet.cell(row=row + r, column=col + c)
#                     cell.border = BORDER
#                     cell.set_explicit_value(str(value))
        self.logger.debug("end fill {0} data".format(sheetName))

    def modifyBorder(self, wb, captureAreas):
        self.logger.debug("start modify border")
        for sname, areas in captureAreas.items():
            sheet = wb[sname]  # 获取sheet表
            for area in areas:
                start, end = area.split(":")
                for row in sheet[start:end]:
                    for cell in row:
                        cell.border = BORDER
        self.logger.debug("end modify border")

    def queryWeb(self, dbsrc, query):
        # 时间戳索引记录
        times = set()
        if not dbsrc:
            self.logger.error("未选择数据源")
            return times, []
        query = Template(query).render(
            DATE_FORMAT=DATE_FORMAT, CURRENT_MONTH=CURRENT_MONTH, CURRENT_TIME=CURRENT_TIME)
        self.logger.debug("queryWeb: query: %s" % query)
        data = {
            "CommandName": "metadata.queryWithDS",
            "Params": {
                "querySQL": query + " FETCH FIRST 1000 ROWS ONLY",
                "dataSourceName": dbsrc
            }
        }
        req = requests.post(URL, json=data, headers=HEADERS)
        result = req.json()
        resultCode = result.get("resultCode", 0)
        resultDesc = result.get("resultDesc", "未知错误")
        self.logger.debug("resultCode: {0}".format(resultCode))
        self.logger.debug("resultDesc: {0}".format(resultDesc))
        if resultCode != 200:
            return times, []

        resultValue = result.get("resultValue", {})
        # 表头
        headers = resultValue.get("columnNames", [])
        for col, cvalue in enumerate(headers):  # 写入表头
            if cvalue.lower().find("time") > -1:
                times.add(col)
        # 数据行
        rows = resultValue.get("rows", [])
        self.logger.debug("rows length: {0}".format(len(rows)))
        return times, rows

    def queryDb(self, dburl, query):
        # 连接数据库查询数据
        engine = create_engine(dburl)
        query = query.replace("%", "%%")
        datas = engine.execute(query).fetchall()
        return datas
