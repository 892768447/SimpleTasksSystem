#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''
Created on 2017年8月14日
@author: Irony."[讽刺]
@site: alyl.vip, orzorz.vip, irony.coding.me , irony.iask.in , mzone.iask.in
@email: 892768447@qq.com
@file: jobs.Reports.FileDaily
@description: 附件邮箱发送公共模块
'''
from datetime import datetime
from email.encoders import encode_base64
from email.header import Header
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
from time import localtime, strftime

from jinja2 import Template
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border, Side
import requests
from sqlalchemy.engine import create_engine

from AutoReportUtil import initLogger  # @UnresolvedImport
# @UnresolvedImport
from AutoReportUtil import sendEmail, DATE_FORMAT, CURRENT_TIME, CURRENT_MONTH


__version__ = "0.0.1"

URL = "http://10.105.4.50:58080/BingoInsight/dataservice.dsr"

HEADERS = {
    "Accept": "*/*",
    "Referer": "http://10.105.4.50:58080/BingoInsight/portal/ClientBin/Bingosoft.DataOne.Portals.Boot.xap",
    "Content-Type": "text/json;chartset=utf-8",
    "User-Agent": "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729; InfoPath.2)",
    "Connection": "Keep-Alive",
    "Cache-Control": "no-cache",
    "Cookie": "JSESSIONID=D639101A895F1F7DC61011877347D8AC",
}

BORDER = Border(
    left=Side("thin", "FF000000"),
    right=Side("thin", "FF000000"),
    top=Side("thin", "FF000000"),
    bottom=Side("thin", "FF000000"),
    vertical=Side("thin", "FF000000"),
    horizontal=Side("thin", "FF000000")
)

MailTemplate = "见附件"


class FileDaily:

    def __init__(self, Id, config, path, subject):
        os.makedirs("datas/logs/", exist_ok=True)
        os.makedirs("datas/tmps/", exist_ok=True)
        self.logger = initLogger("datas/logs/" + Id, Id)
        self.config = config
        self.subject = subject
        filename = os.path.basename(path)  # 文件名
        self.filein = path
        self.fileout = os.path.join("datas/tmps/" + filename)
        self.dataDir = os.path.join(
            "datas/tmps/", os.path.basename(os.path.splitext(self.fileout)[0]))
        os.makedirs(self.dataDir, exist_ok=True)

    def start(self):
        self.logger.debug("pid: %s" % os.getpid())
        # 加载模版文件开始任务
        wb = load_workbook(self.filein)
        # 解析配置
        self.parseConfig(wb)
        # 预发送邮件
        if not self.preSend:
            return self.logger.warn("预发送人员列表为空")
        # 填充数据
        self.fillData(wb)
        # 修改截图区域的边框
#         self.modifyBorder(wb, self.captureArea)
        # 删除多余sheet
        try:
            wb.remove(wb["配置"])
            wb.remove(wb["预发送"])
        except:
            pass
        # 保存填充后的excel
        wb.save(self.fileout)

        # 通过邮箱发送附件
        self._sendEmail(self.subject, self.fileout, self.preSend)
        self.logger.info("send pre email success")

    def _sendEmail(self, subject, file, preSend):
        peoples = preSend  # 如果采用邮箱正式发送
        subject = subject + datetime.now().strftime(" - %Y%m%d %H:%M:%S")
        message = MIMEMultipart()
        message["Subject"] = Header(subject, "gb2312")  # 邮件主题
        # 邮件正文
        content = MIMEText(MailTemplate, _charset="gb2312")
        message.attach(content)
        # 添加附件
        xlsfile = MIMEBase("application", "octet-stream")
        xlsfile.set_payload(open(file, "rb").read())
        xlsfile.add_header("Content-Disposition", "attachment",
                           filename=os.path.basename(file))
        encode_base64(xlsfile)
        message.attach(xlsfile)
        sendEmail(subject, message, peoples)

    def parseConfig(self, wb):
        # 截图区域配置
        self.logger.debug("start parse config")
        self.captureArea = self.config.get("截图", {})
        # 发送人配置
        # 预发送人
        preSheet = wb["预发送"]
        if not preSheet:
            raise Exception("无法找到预发送人表")
        self.preSend = [str(cell.value) for cell in preSheet["A"] if str(
            cell.value) != "None"]  # list(preSheet.columns)[0]]
        self.logger.debug("预发送人: {0}".format(self.preSend))

        # 填充区域配置
        self.fillArea = self.config.get("填充", [])
        self.logger.debug("end parse config")

    def fillData(self, wb):
        self.logger.debug("start fill data")
#         for sheetName, items in self.fillArea.items():
        for item in self.fillArea:
            self.fillItem(wb, item)
        self.logger.debug("end fill data")

    def fillItem(self, wb, item):
        sheetName = item.get("表", None)
        self.logger.warn("item sheetname is none: {0}".format(item))
        if not sheetName:
            return
        sheet = wb[sheetName]  # 获取sheet
        row = item.get("行", 1)
        col = item.get("列", 1)
        typ = item.get("类型", "")
        dbsrc = item.get("数据源", "")
        query = item.get("语句", "")
        self.logger.debug("sheetName:{0}\n"
                          "row:{1}\n"
                          "col:{2}\n"
                          "type:{3}\n"
                          "dbsrc:{4}\n".format(sheetName,
                                               row, col, typ, dbsrc
                                               )
                          )
        if not query:
            return self.logger.warn("查询语句不能为空")
        self.logger.debug("start get {0} data".format(sheetName))
        if typ == "selfsql":
            # 自助报表
            times, datas = self.queryWeb(dbsrc, query)
        else:
            # 连接数据查询
            times, datas = set(), self.queryDb(typ, query)
        self.logger.debug("end get {0} data".format(sheetName))

        self.logger.debug("start fill {0} data".format(sheetName))
        for r, item in enumerate(datas):
            for c, value in enumerate(item):
                if isinstance(value, str) and 1 <= len(value) <= 8:
                    try:
                        value = float(value)
                    except:
                        pass
                try:  # 时间戳转换
                    if c in times:
                        value = localtime(int(str(value)[:10]))
                        value = strftime("%Y-%m-%d %H:%M:%S", value)
                except:
                    pass
                sheet.cell(row=row + r, column=col + c).value = value
#                     cell = sheet.cell(row=row + r, column=col + c)
#                     cell.border = BORDER
#                     cell.set_explicit_value(str(value))
        self.logger.debug("end fill {0} data".format(sheetName))

    def modifyBorder(self, wb, captureAreas):
        self.logger.debug("start modify border")
        for sname, areas in captureAreas.items():
            sheet = wb[sname]  # 获取sheet表
            for area in areas:
                start, end = area.split(":")
                for row in sheet[start:end]:
                    for cell in row:
                        cell.border = BORDER
        self.logger.debug("end modify border")

    def queryWeb(self, dbsrc, query):
        # 时间戳索引记录
        times = set()
        if not dbsrc:
            self.logger.error("未选择数据源")
            return times, []
        query = Template(query).render(
            DATE_FORMAT=DATE_FORMAT, CURRENT_MONTH=CURRENT_MONTH, CURRENT_TIME=CURRENT_TIME)
        self.logger.debug("queryWeb: query: %s" % query)
        data = {
            "CommandName": "metadata.queryWithDS",
            "Params": {
                "querySQL": query + " FETCH FIRST 1000 ROWS ONLY",
                "dataSourceName": dbsrc
            }
        }
        req = requests.post(URL, json=data, headers=HEADERS)
        result = req.json()
        resultCode = result.get("resultCode", 0)
        resultDesc = result.get("resultDesc", "未知错误")
        self.logger.debug("resultCode: {0}".format(resultCode))
        self.logger.debug("resultDesc: {0}".format(resultDesc))
        if resultCode != 200:
            return times, []

        resultValue = result.get("resultValue", {})
        # 表头
        headers = resultValue.get("columnNames", [])
        for col, cvalue in enumerate(headers):  # 写入表头
            if cvalue.lower().find("time") > -1:
                times.add(col)
        # 数据行
        rows = resultValue.get("rows", [])
        self.logger.debug("rows length: {0}".format(len(rows)))
        return times, rows

    def queryDb(self, dburl, query):
        # 连接数据库查询数据
        engine = create_engine(dburl)
        query = query.replace("%", "%%")
        datas = engine.execute(query).fetchall()
        return datas
