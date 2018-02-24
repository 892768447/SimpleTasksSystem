#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''
Created on 2017年7月17日
@author: Irony."[讽刺]
@site: alyl.vip, orzorz.vip, irony.coding.me , irony.iask.in , mzone.iask.in
@email: 892768447@qq.com
@file: jobs.MmsCollectionJob.MmsCollection
@description: 
'''

__version__ = "0.0.1"

from concurrent.futures.thread import ThreadPoolExecutor
from datetime import datetime, timedelta
import encodings.idna  # @UnusedImport
import json
import os
from queue import Queue
import threading
from time import sleep
import traceback

from openpyxl.reader.excel import load_workbook
import pymysql  # @UnusedImport
import requests
from sqlalchemy.dialects.mysql.types import DOUBLE, VARCHAR, INTEGER, DATETIME, TIMESTAMP
from sqlalchemy.engine import create_engine
from sqlalchemy.ext.declarative.api import declarative_base
from sqlalchemy.orm.scoping import scoped_session
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.pool import QueuePool
from sqlalchemy.sql.expression import and_, text
from sqlalchemy.sql.schema import Column
import yaml

from AutoReportUtil import initLogger  # @UnresolvedImport
from AutoReportUtil import sendEmailText  # @UnresolvedImport


HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/7.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET4.0C; .NET4.0E; InfoPath.3)"
}

DirPath = "datas/tmps/mmscollection"
os.makedirs(DirPath, exist_ok=True)
YEAR = datetime.now().year

Base = declarative_base()
Class_Dict = {}

class AbstractClass(Base):

    __abstract__ = True

# 数据库表基类,后续为动态生成
class MysqlBasKpiInfos(object):  # 数据库表结构

    # __tablename__ = "baskpiinfos"

    # ['F_TODAY_VALUE', 'F_LAST_MONTH_VALUE', 'N_DAY',
    # 'AREA_CODE_AS', 'F_CHAIN_RAT', 'F_YEAR_RAT',
    # 'F_MONTH_ACC_VALUE', 'F_YEAR_ACC_VALUE', 'AREA_NAME']

    # 自增id
    ID = Column(INTEGER, primary_key=True)
    # 采集日期
    CollectDate = Column(DATETIME, nullable=False)
    # 查询日期
    InsertDate = Column(TIMESTAMP, server_default=text(
        "CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP"))  # 插入时间
    # 指标代码
    BasKpiCode = Column(VARCHAR(255), nullable=False)
    # 区域代码
    BasAreaCode = Column(INTEGER, nullable=False)  # AREA_CODE_AS--3
    # 今天的值
    TodayValue = Column(DOUBLE, default=0)  # F_TODAY_VALUE--0
    # 上周的值
    LastWeekDayValue = Column(DOUBLE, default=0)  #
    # 上个月的值
    LastMonthDayValue = Column(DOUBLE, default=0)  # F_LAST_MONTH_VALUE--1
    # 去年的值
    LastYearDayValue = Column(DOUBLE, default=0)
    # 月值
    MonthTotalValue = Column(DOUBLE, default=0)  # F_MONTH_ACC_VALUE--6
    # 年总值
    YearTotalValue = Column(DOUBLE, default=0)  # F_YEAR_ACC_VALUE--7
    # 今天预定量
    TodayOrder = Column(DOUBLE, default=0)

def get_class_name_and_table_name(prefixname):
    name = "baskpiinfos" + str(prefixname)
    return name, name

def getClass(prefix, hashid):
    prefixname = str(prefix) + str(hashid)
    if prefixname not in Class_Dict:
        class_name, table_name = get_class_name_and_table_name(prefixname)
        cls = type(
            class_name,
            (AbstractClass, MysqlBasKpiInfos),
            {"__tablename__": table_name}
        )
        Class_Dict[prefixname] = cls
        return cls
    return Class_Dict[prefixname]

class MmsCollection:

    def __init__(self, Id):
        self.Id = Id
        self.data_queue = Queue()
        self._start_time = datetime.now()
        self.logger = initLogger("datas/logs/" + Id, Id)

    def getData(self, start, end, areaCode, kpiCode, year, clazz, twice=1):
        # 获取数据
        postdata = json.dumps(
            {"CommandName": "GetDayTimeTrendLast",
             "Params": {"StartDate": "%s" % start,
                        "EndDate": "%s" % end,
                        "AreaCodeAsList": "%s" % areaCode,
                        "IndexID": "%s" % kpiCode,
                        "IndexMarketType": "",
                        "Year": "%s" % year}
             }
        )
        # print(postdata)
        try:
            req = requests.post(url=self.url,
                                data=postdata,
                                headers=HEADERS,
                                verify="cacert.pem",
                                timeout=90)
            body = req.content.decode(req.encoding).replace("'", '"')
            try:
                body = json.loads(body)
                if not body.get("rowCount", None):
                    return  # 没有数据跳过
                rows = body.get("rows")
            except Exception as e:
                    rows = None
            if rows:
                self.logger.debug(
                    "%s,已完成一条查询：%s,,%s" % (threading.currentThread(), areaCode, kpiCode))
                for row in rows:
                    try:  # 由于这里多加了一个字段导致错误
                        TodayValue, LastMonthDayValue, CollectDate, _, BasAreaCode, F_CHAIN_RAT, F_YEAR_RAT, MonthTotalValue, YearTotalValue, AREA_NAME = row  # @UnusedVariable
                    except:
                        try:  # 原始没有多加字段的情况
                            TodayValue, LastMonthDayValue, CollectDate, BasAreaCode, F_CHAIN_RAT, F_YEAR_RAT, MonthTotalValue, YearTotalValue, AREA_NAME = row  # @UnusedVariable
                        except:
                            self.logger.error(
                    "%s,任务（%s-%s-%s-%s-%s）字段解析失败: %s" % (threading.currentThread(), start, end, areaCode, kpiCode, year, e.splitlines()[-1]))
                            continue
                    if not TodayValue:
                        TodayValue = 0
                    if not LastMonthDayValue:
                        LastMonthDayValue = 0
                    if not MonthTotalValue:
                        MonthTotalValue = 0
                    if not YearTotalValue:
                        YearTotalValue = 0
#                     if self.ispatch:
#                         method = dict
#                     else:
#                         method = clazz
#                     Infos = method(
                    Infos = dict(
                        CollectDate=CollectDate,
#                         BasDate=datetime.now().strftime("%Y%m%d %H:%M:%S"),
                        BasKpiCode=kpiCode,
                        BasAreaCode=BasAreaCode,
                        TodayValue=TodayValue,
                        LastWeekDayValue=0,
                        LastMonthDayValue=LastMonthDayValue,
                        LastYearDayValue=0,
                        MonthTotalValue=MonthTotalValue,
                        YearTotalValue=YearTotalValue,
                        TodayOrder=0
                    )
                    self.data_queue.put_nowait((Infos, clazz))  # 放入队列数据
            else:
                self.logger.info(
                    "%s,已完成一条查询：%s,%s-%s %s %s,但是没有数据" % (threading.currentThread(),
                        start, end, areaCode, kpiCode, str(req)))
        except Exception as e:
            e = traceback.format_exc()
            if twice < 4:
                twice += 1
                self.logger.error("%s,任务（%s-%s-%s-%s-%s）重试第%d次: %s" % 
                                  (threading.currentThread(), start, end, areaCode, kpiCode, year, twice, e.splitlines()[-1]))
                # sleep(self.sleepnum * (2 ** twice))  # 暂停时间
                self.executor.submit(
                    self.getData, start, end, areaCode, kpiCode, year, clazz, twice)
            else:
                self.logger.error(
                    "%s,任务（%s-%s-%s-%s-%s）失败: %s" % (threading.currentThread(), start, end, areaCode, kpiCode, year, e.splitlines()[-1]))
        sleep(self.sleepnum)  # 抓取一个后休息

    def _insert(self, dates):
#         print("总共抓取：", self.data_queue.qsize())
        self.logger.debug("总共抓取：%d" % self.data_queue.qsize())
        if self.data_queue.empty():
            return
        tdatas = [self.data_queue.get_nowait() for _ in range(self.data_queue.qsize())]
        self._insertAll(dates, tdatas)
    
    def _insertAll(self, dates, tdatas):
        session = self.session()
        for start, end, year, clazz in dates:  # @UnusedVariable
            st = datetime.strptime(start, "%Y%m%d").strftime("%Y-%m-%d")
            en = datetime.strptime(end, "%Y%m%d").strftime("%Y-%m-%d")
            self.logger.debug("开始清空数据: %s, %s" % (st, en))
            sql = session.query(clazz).filter(
                and_(
                    clazz.CollectDate >= st,
                    clazz.CollectDate <= en,
                    clazz.BasKpiCode.in_(self.kpi_code),
                    clazz.BasAreaCode.in_(self.area_code)
                )
            )  # 满足指定日期范围并且在指定的指标和区域中的数据进行删除
            sql.delete(synchronize_session=False)
            # print(str(sql), "  :  ", sql.delete(synchronize_session=False))
        session.commit()
        session.close()
#         print("开始导入")
        self.logger.debug("开始导入数据")
        adatas = {}
        for Infos, clazz in tdatas:
            if clazz not in adatas:
                adatas[clazz] = []
            adatas[clazz].append(Infos)
        session = self.session()
        for clazz in adatas.keys():
            session.execute(
                clazz.__table__.insert(),
                adatas.get(clazz, [])
            )
        session.commit()  # 提交到数据库中
        session.close()
    
    def _startTasks(self, tname, dates):
        # 把所有的任务放进队列中
        _tasks = set()
        for kpiCode in self.kpi_code:  # kpi指标
            for areaCode in self.area_code:  # 区域码
                for startdate, enddate, year, clazz in dates:  # 抓取日期
                    _tasks.add((startdate, enddate, areaCode, kpiCode, year, clazz))
        self.logger.info("总共 %d 个任务" % len(_tasks))
#         print("总共 %d 个任务" % len(_tasks))
        for index, task in enumerate(_tasks):  # @UnusedVariable
            self.executor.submit(self.getData, *task)
#         print("任务分发完毕,等待抓取完成")
        self.logger.debug("任务分发完毕,等待抓取完成")
        self.executor.shutdown(wait=True)  # 等待所有任务完成
        '''
        if self.ispatch:  # 补数据需要批量
            print("补数据批量插入")
            size = self.data_queue.qsize()
            #print("开始数据数量对比")
            #self.logger.info("开始数据数量对比")
#             can, oldCount = self.compareData(dates, size)
#             print(can, oldCount, size)
#             if not can:  # 对比数据
#                 print("数量匹配有误，取消导入", oldCount, size)
#                 self.logger.error("数量匹配有误，取消导入：oldCount：%s，newCount：%s" % (oldCount, size))
#                 self.logger.info("任务完成,用时%s" % 
#                          str(datetime.now() - self._start_time))
#                 return
            num = int((size - size % 10000) / 10000)
            self.executors = ThreadPoolExecutor(max_workers=num + 1)
            tdatas = []
            for _ in range(num):
                tdatas.clear()
                for __ in range(10000):
                    try: tdatas.append(self.data_queue.get_nowait())
                    except: pass
                self.executors.submit(self._insertAll, dates, tdatas.copy())
            tdatas.clear()
            for __ in range(size % 10000):
                try: tdatas.append(self.data_queue.get_nowait())
                except: pass
            self.executors.submit(self._insertAll, dates, tdatas.copy())
            tdatas.clear()
            self.executors.shutdown(wait=True)  # 等待所有任务完成
        else:
            print("批量添加")
            self._insert(dates)
        '''
#         print("批量添加")
        self.logger.debug("批量添加")
        self._insert(dates)
#         print("任务已完成")
        self.logger.info("抓取任务已完成,用时%s" % 
                         str(datetime.now() - self._start_time))
        # 发送邮件通知
        self.logger.debug(str(sendEmailText(self.preSend, self.Id, "任务:%s 已完成" % self.Id)))

    def compareData(self, dates, size):
        # 对比数据
        oldCount = 0
#             tmp_queue = [item for item in self.data_queue.queue]#取出所有结果
        session = self.session()
        for start, end, year, clazz in dates:  # @UnusedVariable
            st = datetime.strptime(start, "%Y%m%d").strftime("%Y-%m-%d")
            en = datetime.strptime(end, "%Y%m%d").strftime("%Y-%m-%d")
            sql = session.query(clazz).filter(
                and_(
                    clazz.CollectDate >= st,
                    clazz.CollectDate <= en,
                    clazz.BasKpiCode.in_(self.kpi_code),
                    clazz.BasAreaCode.in_(self.area_code)
                )
            )
            oldCount += sql.count()
        session.close()
        if oldCount == 0:
            return 1, oldCount
        return abs(oldCount - size) < self.comparenum, oldCount

    def _getDates(self, start, end):
        # 处理跨年
        if start.year < end.year:
            dates = [
                (start.strftime("%Y%m%d"), datetime(
                    start.year, 12, 31).strftime("%Y%m%d"), str(start.year),
                 getClass(self.prefixname, start.year))
            ]
            for year in range(1, end.year - start.year):
                _year = year + start.year
                dates.append(
                    (datetime(_year, 1, 1).strftime(
                    "%Y%m%d"), datetime(_year, 12, 31).strftime("%Y%m%d"),
                    str(_year), getClass(self.prefixname, _year))
                )
            dates.append(
                (datetime(end.year, 1, 1).strftime("%Y%m%d"),
                          end.strftime("%Y%m%d"), str(end.year),
                          getClass(self.prefixname, end.year))
            )
        else:  # 没有跨年
            dates = [
                (start.strftime("%Y%m%d"),
                 end.strftime("%Y%m%d"), str(end.year), getClass(self.prefixname, end.year))
            ]
#         print(dates)
        self.logger.debug(dates)
        self.logger.info("抓取如下期间的数据：%s" % str(dates))
        return dates

    def start(self, path):
        now = datetime.now()
        self.logger.info("**********%s**********" % 
                         now.strftime("%Y-%m-%d %H-%M-%S"))
        
        wb = load_workbook(path)  # 加载excel文件

        config = yaml.load(wb["配置文件"].cell(row=1, column=1).value)
#         config = json.load(open("配置文件.json", encoding="utf-8"))

        # 接口地址
        self.url = config.get("网址", None)
        if not self.url:
            self.logger.error("配置文件 中网址为空，程序退出。")
            return

        self.sleepnum = config.get("休眠时间", 2)
        self.ispatch = config.get("补数据", None) == "是"
        self.timeslot = config.get("时间段", [])
        self.prefixname = config.get("自定义扩展表名", "")

        # 加载区域代码
        self.area_code = [str(cell.value) for cell in wb["区域代码"]["A"] if str(cell.value) != "None"]
#         self.area_code = json.load(
#             open("区域代码.json", encoding="utf-8")).get("区域代码", None)
#         print(len(self.area_code), self.area_code)
        self.logger.debug("area len: {0}, areas: {1}".format(len(self.area_code), self.area_code))
        if not self.area_code:
            self.logger.error("区域代码加载失败，请检查")
            return

        # 加载指标代码
        self.kpi_code = [str(cell.value) for cell in wb["指标代码"]["A"] if str(cell.value) != "None"]
#         self.kpi_code = json.load(
#             open("指标代码.json", encoding="utf-8")).get("指标代码", None)
#         print(len(self.kpi_code), self.kpi_code)
        self.logger.debug("kpi len: {0}, kpis: {1}".format(len(self.kpi_code), self.kpi_code))
        if not self.kpi_code:
            self.logger.error("指标代码加载失败，请检查")
            return
        
        # 数据库地址
        databaseurl = config.get("数据库连接地址", None)
        if not databaseurl:
#             print("没有找到数据库连接地址")
            self.logger.error("没有找到数据库连接地址")
            return
        
        # 采集天数
        self.days = config.get("天数", 10)

        if self.ispatch and isinstance(self.timeslot, list) and len(self.timeslot) == 2:
            start = datetime.strptime(self.timeslot[0], "%Y-%m-%d")
            end = datetime.strptime(self.timeslot[1], "%Y-%m-%d")
            if start.year != end.year:
#                 print("起始时间段必须在同一年")
                self.logger.error("起始时间段必须在同一年")
                return
            # 保证最小的日期在前面
            if start > end:
                start, end = end, start
        else:  # 当前日期的前self.days天
            start = now + timedelta(days=-self.days)
            end = now + timedelta(days=-1)
        
        self.comparenum = config.get("比对差量", 100)
        self.threadnum = config.get("线程数", 4)
        self.executor = ThreadPoolExecutor(max_workers=self.threadnum)
        self.logger.info("线程数量: %s" % str(self.threadnum))
        
        _dates = self._getDates(start, end)  # 日期段

        try:
            engine = create_engine(
                databaseurl,
                pool_size=self.threadnum,
                poolclass=QueuePool,
            )
            Base.metadata.create_all(engine)  # 创建表
            self.session = scoped_session(
                sessionmaker(bind=engine,
                             autocommit=False,
                             autoflush=True
                )
            )
        except Exception as e:
            traceback.print_exc()  # 初始化失败
            self.logger.error(str(e))
            return
        
        # 预发送人
        preSheet = wb["预发送"]
        if not preSheet:
            self.logger.warn("无法找到预发送人表")
            self.preSend = []
        else:
            self.preSend = [str(cell.value) for cell in preSheet["A"] if str(cell.value) != "None"]  # list(preSheet.columns)[0]]
        self.logger.debug("预发送人: {0}".format(self.preSend))
        self._startTasks(path, _dates)
