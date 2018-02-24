#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''
Created on 2017年7月18日
@author: Irony."[讽刺]
@site: alyl.vip, orzorz.vip, irony.coding.me , irony.iask.in , mzone.iask.in
@email: 892768447@qq.com
@file: jobs.ReportCollectionJob.BaseCollection
@description: 
'''
from datetime import datetime, timedelta
import os
from random import random, randint
import re
from time import sleep
import traceback

from openpyxl.reader.excel import load_workbook
import requests
from sqlalchemy.engine import create_engine
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.sql.expression import and_

from AutoReportUtil import sendEmailText  # @UnresolvedImport


# from memory_profiler import profile
# from bs4 import BeautifulSoup
# from xlwt.Formatting import Pattern
# from xlwt.Style import XFStyle
# from xlwt.Workbook import Workbook
__version__ = "0.0.1"

ThPattern = re.compile("<TH.*?>(.*?)</TH>")
# TdPattern = re.compile("<tdclass='txt'>(.*?)</td>")
TdPattern = re.compile("<td class='txt'>(.*?)</td>")


class BaseCollection:

    Name = ""  # 名字
    TmpDir = "datas/tmps"  # tmp目录
    DataDir = "datas"  # data目录

    def getModel(self):
        pass

    def getBase(self):
        pass

    def getTime(self):
        pass

    def __init__(self, logger, config, codes, headers, cookies, referers, randomnum, urls, params, tname):
        self.logger = logger
        os.makedirs(self.TmpDir, exist_ok=True)
#         os.makedirs(self.DataDir, exist_ok=True)
        self.randomnum = randomnum
        self.Codes = codes
        self.Headers = headers
        self.Config = config
        self.tname = tname
        wb = load_workbook(tname)
        # 预发送人
        preSheet = wb["预发送"]
        if not preSheet:
            self.logger.warn("无法找到预发送人表")
            self.preSend = []
        else:
            self.preSend = [str(cell.value) for cell in preSheet["A"] if str(
                cell.value) != "None"]  # list(preSheet.columns)[0]]
        self.logger.debug("预发送人: {0}".format(self.preSend))

        self._referer = referers.get(self.Name, "")
        if not self._referer:
            self.logger.error("读取{0}Referer信息错误".format(self.Name))
            raise Exception("读取{0}Referer信息错误".format(self.Name))

        cookie = cookies.get("报表", "")
        if not cookie:
            self.logger.error("读取{0}Cookie信息错误".format(self.Name))
            raise Exception("读取{0}Cookie信息错误".format(self.Name))
        self.Headers["Cookie"] = cookie.format(random=randomnum)

        self.Url = urls.get(self.Name, None)
        if not self.Url:
            self.logger.error("读取{0}网址错误".format(self.Name))
            raise Exception("读取{0}网址错误".format(self.Name))

        self.Param = params.get(self.Name, None)
        if not self.Param:
            self.logger.error("读取{0}表单错误".format(self.Name))
            raise Exception("读取{0}表单错误".format(self.Name))

    def delOldData(self, session):
        # 删除旧数据
        try:
            now = datetime.now()
#             print("delOldData now: ", now)
#             self.logger.info("delOldData now: %s" % now)
#             st = self.Config.get("start", datetime.now().strftime("%Y%m%d"))
#             en = self.Config.get("end", datetime.now().strftime("%Y%m%d"))
#             st = datetime.strptime(st, "%Y%m%d").strftime("%Y-%m-%d 00:00:00")
#             en = datetime.strptime(en, "%Y%m%d").strftime("%Y-%m-%d 23:59:59")
            if now.day == 1:  # 月初
                st = now.strftime("%Y-%m-%d 00:00:00")
            else:
                st = (now + timedelta(days=self.Config.get("day", 0))
                      ).strftime("%Y-%m-%d 00:00:00")
            en = now.strftime("%Y-%m-%d 23:59:59")
#             print("del old data: ", st, en)
            self.logger.info("del old data: %s %s" % (st, en))
            session.query(self.getModel()).filter(
                and_(self.getTime() >= st, self.getTime() <= en)
            ).delete()
            session.commit()
        except Exception as e:
            #             print(e)
            self.logger.warn(str(e))

    def getParam(self, index):
        # 提取数据
        now = datetime.now()
#         print("getParam now: ", now)
#         self.logger.info("getParam now: %s" % now)
#         date = now.strftime("%Y%m%d")
#         print("date: %s" % date)
#         self.logger.info("date: %s" % date)
#         datestart = self.Config.get("start", datetime.now().strftime("%Y%m%d"))
#         dateend = self.Config.get("end", datetime.now().strftime("%Y%m%d"))
        if now.day == 1:  # 月初
            datestart = now.strftime("%Y%m%d")
        else:
            datestart = (now + timedelta(days=self.Config.get("day", 0))
                         ).strftime("%Y%m%d")
        dateend = now.strftime("%Y%m%d")
#         print("getParam date: ", datestart, dateend)
        self.logger.debug("getParam date: %s %s" % (datestart, dateend))
        now = datetime.now().strftime("%Y%m%d+%H%%3A%M%%3A%S")
#         print("getParam now : %s" % now)
        self.logger.info("getParam now : %s" % now)
        param = self.Param.format(
            index=index, datestart=datestart, dateend=dateend, now=now)
        return param.encode()

#     def parseHeader(self, sheet, table):
#         # 解析表头
#         ths = table.find_all("th")
#         _ths = [th.getText() for th in ths]
#         # 写入表头
#         style = XFStyle()  # 样式
#         pattern = Pattern()
#         pattern.pattern = Pattern.SOLID_PATTERN
#         pattern.pattern_fore_colour = 0x16
#         style.pattern = pattern
#         for col, header in enumerate(_ths):
#             sheet.write(0, col, header, style)
#         # self.logger.info("headers: %s"%_ths)
#         return _ths

    '''
    @profile
    def parseBody(self, sheet, table):
        # 解析内容
        trs = table.find_all("tr")
        result = []
        for row, tr in enumerate(trs):  # @UnusedVariable
            tds = tr.find_all("td")
            _tds = [td.getText() for td in tds]
            if not _tds[0].startswith("合计") and len(_tds[0]) > 0:
                result.append(_tds)
            del tds
            del _tds
#             # 写入内容
#             for col, text in enumerate(_tds):
#                 if col == 0 and text.find("合计") > -1:
#                     continue
#                 sheet.write(row + 1, col, text)
        # self.logger.info(str(result[-2:]))
        del trs
        return result

    @profile
    def parser(self, name, data):
        # 解析数据
        bs = BeautifulSoup(data, "lxml")
        del data
        table = bs.find_all("table")
        bs.decompose()  # 释放?
        bs = None
        del bs  # 找到table后就删除bs,防止内存过高
#         print("table len: %s" % len(table))
        self.logger.info("table len: %s" % len(table))
        if len(table) != 4:
            return
#         wbk = Workbook()  # 表格
#         sheet = wbk.add_sheet(os.path.basename(name), cell_overwrite_ok=True)
        sheet = None
#         self.parseHeader(sheet, table[2])  # 表头
        bodys = self.parseBody(sheet, table[3])  # 数据
#         sheet.set_panes_frozen(True)
#         sheet.set_horz_split_pos(1)  # 冻结表头
#         wbk.save(name)  # 保存
#         self.logger.info("%s %s"%(len(headers), len(bodys)))
        del table
        return bodys
    '''

#     @profile
    def parser(self, name, data):
        try:
            #             data = data.replace(" ", "")  # 替换空格
            # 先对数据进行查找
            ti1 = data.find("<table><tr>")
#             ti2 = data.find("</table>", ti1)
            if ti1 < 0:  # or ti2 < 0:
                self.logger.error("解析数据时未找到关键点")
                return []
            data1 = data[0:ti1]  # 包含表头的数据 @UnusedVariable
#             data2 = data[ti1:ti2].split("</tr>")  # 包含数据
            data2 = data[ti1:].split("</tr>")  # 包含数据
            del data
            # 匹配表头
#             print("表头:", ThPattern.findall(data1))
            # 匹配字段
            results = (TdPattern.findall(d) for d in data2 if d)
            # 替换统计
#             results = [d for d in results if d and d[0].find("合计") == -1]
            # 替换两边空格和统计
            results = ((i.strip() for i in d)
                       for d in results if d and d[0].find("合计") == -1)
            del data1, data2
            return results if results else ()
#             return results if results[0:-3] else []
        except Exception as e:
            self.logger.error("parser error: %s" % e)
        return ()

#     @profile
    def run(self, host):
        try:
            # 随机休眠
            self.Headers["Referer"] = self._referer.format(
                host=host, random=self.randomnum)
            self.Url = self.Url.format(
                host=host, random=self.randomnum, rnd=random())
            sleep(randint(3, 6))
            # 数据库连接
            _model = self.getModel()
            ext = "_" + datetime.now().strftime("%Y%m") if self.Config.get("SPLIT", 0) else ""
            setattr(_model, "__tablename__", _model.__tablename__ + ext)
            print("table name: " + self.getModel().__tablename__)
            self.Config["dbname"] = self.getModel().__dbname__
            self.logger.info(str(self.Config))
    #         print(self.Config)
            try:
                # 初始化数据库连接:
                #             print("mysql+pymysql://{user}:{pwd}@{host}".format(**self.Config))
                engine = create_engine(
                    "mysql+pymysql://{user}:{pwd}@{host}".format(**self.Config))
                # 创建数据库
                with engine.connect() as con:
                    con.execute(
                        "CREATE DATABASE IF NOT EXISTS {0} CHARACTER SET utf8;".format(
                            self.getModel().__dbname__))
            except Exception as e:
                #             print("创建数据库错误 %s %s" % (self.getModel().__dbname__, e))
                self.logger.warn("创建数据库错误 %s %s" %
                                 (self.getModel().__dbname__, e))
            try:
                self.engine = create_engine(
                    "mysql+pymysql://{user}:{pwd}@{host}/{dbname}?charset=utf8".format(**self.Config))
                self.getBase().metadata.create_all(self.engine)  # 创建库和表
            except Exception as e:
                #             print("初始数据库和表失败:%s %s %s" % (self.getModel().__dbname__, self.getModel().__tablename__, e))
                self.logger.warn("初始数据库和表失败:%s %s %s" % (
                    self.getModel().__dbname__, self.getModel().__tablename__, e))
                raise Exception(e)

            allDatas = []
            now = datetime.now().strftime("%Y-%m-%d")  # %H-%M-%S
            for name, index in self.Codes.items():
                self.logger.debug("get %s %s" % (name, index))
                try:
                    req = requests.post(url=self.Url,
                                        data=self.getParam(index), headers=self.Headers
                                        )
                    cdes = req.headers.get("Content-Disposition", None)
                    if not cdes or cdes.find("attachment") == -1:
                        #                     print("{0}: 采集{1}-{2}失败:登录已失效,{3}".format(self.Name, index, name, req.headers))
                        self.logger.warn(
                            "{0}: 采集{1}-{2}失败:登录已失效,{3}".format(self.Name, index, name, req.headers))
                        self.logger.debug(req.content)
                        open(self.TmpDir + "/" + self.Name + "error.log",
                             "w").write("{0}: 采集{1}-{2}失败:登录已失效".format(self.Name, index, name))
                    os.makedirs(self.TmpDir + "/" + now, exist_ok=True)
    #                 os.makedirs(self.DataDir + "/" + now, exist_ok=True)
                    # 把抓取的数据保存为文件
                    open(self.TmpDir + "/%s/%s.xls" %
                         (now, name), "wb").write(req.content)
                    self.logger.debug("save %s/%s/%s.xls ok" %
                                      (self.TmpDir, now, name))
                    # 解析数据生成xls并入库
                    req.encoding = "gbk"
                    self.logger.debug("start parser %s %s" % (name, index))
                    bodys = self.parser(self.DataDir + "/%s/%s.xls" % (now, name),
                                        req.text)
                    del req  # 删除这个请求
                    if bodys:
                        allDatas.extend(bodys)
                    self.logger.debug("end parser %s %s" % (name, index))
                except Exception as e:
                    #                 print("{0}: 采集{1}-{2}失败:{3}".format(self.Name, index, name, e))
                    self.logger.warn(
                        "{0}: 采集{1}-{2}失败:{3}".format(self.Name, index, name, e))
                    traceback.print_exc()
                    return
            if not allDatas:
                return self.logger.warn("数据为空")
            # 开始入库
            session = sessionmaker(bind=self.engine)()
            # 删除当天旧数据
    #         print("删除 %s %s" % (self.Name, "当天旧数据"))
            self.logger.info("删除 %s %s" % (self.Name, "当天旧数据"))
            self.delOldData(session)
    #         print("开始入库：%s" % self.Name)
            self.logger.info("开始入库：%s" % self.Name)
            # 批量插入
            _allDatas = [dict(zip(self.getModel().keys(), value))
                         for value in allDatas]
            del allDatas  # 删除,防止内存过高
            session.execute(
                self.getModel().__table__.insert(),
                _allDatas
            )
            session.commit()
            session.close()
            del _allDatas  # 删除,防止内存过高
    #         print("入库完成：%s" % self.Name)
            # 发送邮件通知
            # 发送邮件通知
            self.logger.debug(
                str(sendEmailText(self.preSend, self.Name, "任务:%s 已完成" % self.Name)))
            self.logger.info("入库完成：%s" % self.Name)
        except:
            self.logger.error(traceback.format_exc())
